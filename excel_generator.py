import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from pathlib import Path
from typing import List, Dict, Any

# Configurações de Cores Premium (Paleta Classic Navy)
NAVY_HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
ICE_BLUE_FILL = PatternFill(start_color="F2F5F9", end_color="F2F5F9", fill_type="solid")
WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
ACCENT_BLUE_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

FONT_NAME = "Segoe UI"
FONT_TITLE = Font(name=FONT_NAME, size=16, bold=True, color="1F4E78")
FONT_SUBTITLE = Font(name=FONT_NAME, size=11, italic=True, color="595959")
FONT_HEADER = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")
FONT_BODY = Font(name=FONT_NAME, size=10)
FONT_BODY_BOLD = Font(name=FONT_NAME, size=10, bold=True)
FONT_LINK = Font(name=FONT_NAME, size=10, color="0563C1", underline="single")

BORDER_THIN = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)

BORDER_BOTTOM_DOUBLE = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='double', color='1F4E78')
)

ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=True)

def generate_excel_report(
    uasg: str,
    num_compra: str,
    objeto: str,
    propostas_analisadas: List[Dict[str, Any]],
    output_path: Path
):
    """
    Gera a planilha de resultados do Analisador de Propostas Comercial.
    Aba 1: Resumo do Pregão
    Aba 2: Tabela de Confronto das Propostas
    """
    wb = openpyxl.Workbook()
    
    # =====================================================================
    # ABA 1: RESUMO DO PREGÃO
    # =====================================================================
    ws_resumo = wb.active
    ws_resumo.title = "Resumo"
    ws_resumo.views.sheetView[0].showGridLines = True
    
    # Título Principal
    ws_resumo["B2"] = "Relatório de Análise de Propostas"
    ws_resumo["B2"].font = FONT_TITLE
    ws_resumo["B3"] = "Consolidação automática via Inteligência Artificial"
    ws_resumo["B3"].font = FONT_SUBTITLE
    
    # Bloco de Informações Gerais
    info_labels = [
        ("UASG", uasg),
        ("Número da Compra", num_compra),
        ("Objeto", objeto)
    ]
    
    current_row = 5
    for label, val in info_labels:
        cell_label = ws_resumo.cell(row=current_row, column=2, value=label)
        cell_label.font = FONT_BODY_BOLD
        cell_label.fill = ACCENT_BLUE_FILL
        cell_label.alignment = Alignment(horizontal="left", vertical="top")
        cell_label.border = BORDER_THIN
        
        cell_val = ws_resumo.cell(row=current_row, column=3, value=val)
        cell_val.font = FONT_BODY
        cell_val.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        cell_val.border = BORDER_THIN
        
        if label == "Objeto":
            ws_resumo.row_dimensions[current_row].height = 60
            ws_resumo.merge_cells(start_row=current_row, start_column=3, end_row=current_row, end_column=5)
            # Aplica borda nas células mescladas
            for col in range(3, 6):
                ws_resumo.cell(row=current_row, column=col).border = BORDER_THIN
        else:
            ws_resumo.row_dimensions[current_row].height = 24
            ws_resumo.merge_cells(start_row=current_row, start_column=3, end_row=current_row, end_column=5)
            for col in range(3, 6):
                ws_resumo.cell(row=current_row, column=col).border = BORDER_THIN
        
        current_row += 1
        
    # Ajuste de largura das colunas da aba de Resumo
    ws_resumo.column_dimensions['A'].width = 3
    ws_resumo.column_dimensions['B'].width = 22
    ws_resumo.column_dimensions['C'].width = 25
    ws_resumo.column_dimensions['D'].width = 25
    ws_resumo.column_dimensions['E'].width = 30
    
    # =====================================================================
    # ABA 2: CONFRONTO DE PROPOSTAS
    # =====================================================================
    ws_confronto = wb.create_sheet(title="Análise de Propostas")
    ws_confronto.views.sheetView[0].showGridLines = True
    
    headers = [
        "Fornecedor\n(Automático)",
        "CNPJ\n(Automático)",
        "PDF da Proposta\n(Link)",
        "Item TR\n(Automático)",
        "Item Ofertado\n(Automático)",
        "Confronto IA\n(Automático)",
        "Setor Competente\n(Manual)",
        "Parecer Técnico\n(Dropdown)",
        "Justificativa\n(Manual)"
    ]
    
    # Configura os cabeçalhos
    ws_confronto.row_dimensions[1].height = 36
    for idx, header in enumerate(headers, 1):
        cell = ws_confronto.cell(row=1, column=idx, value=header)
        cell.font = FONT_HEADER
        cell.fill = NAVY_HEADER_FILL
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_BOTTOM_DOUBLE
        
    # Popula dados das propostas
    row_num = 2
    for prop in propostas_analisadas:
        for item in prop["itens_analisados"]:
            ws_confronto.row_dimensions[row_num].height = 80
            
            # 1. Fornecedor
            c1 = ws_confronto.cell(row=row_num, column=1, value=prop["razao_social"])
            c1.font = FONT_BODY
            c1.alignment = ALIGN_LEFT
            c1.border = BORDER_THIN
            
            # 2. CNPJ
            c2 = ws_confronto.cell(row=row_num, column=2, value=prop["cnpj"])
            c2.font = FONT_BODY
            c2.alignment = ALIGN_CENTER
            c2.border = BORDER_THIN
            
            # 3. PDF da Proposta (Hiperlink local relativo)
            relative_pdf_path = prop["relative_pdf_path"]
            filename = Path(relative_pdf_path).name
            # Utiliza a fórmula de HIPERLINK do Excel
            # A fórmula assume que a planilha está na raiz e a pasta de propostas é relativa
            c3 = ws_confronto.cell(row=row_num, column=3)
            c3.value = f'=HYPERLINK("{relative_pdf_path}", "Abrir PDF ({filename})")'
            c3.font = FONT_LINK
            c3.alignment = ALIGN_CENTER
            c3.border = BORDER_THIN
            
            # 4. Item TR
            item_tr_desc = f"Item {item['item_tr_num']}:\n{item['item_tr_nome']}"
            c4 = ws_confronto.cell(row=row_num, column=4, value=item_tr_desc)
            c4.font = FONT_BODY
            c4.alignment = ALIGN_LEFT
            c4.border = BORDER_THIN
            
            # 5. Item Ofertado
            item_ofertado_desc = (
                f"Desc: {item['descricao']}\n"
                f"Marca: {item['marca'] or 'Não informada'}\n"
                f"Modelo: {item['modelo'] or 'Não informado'}\n"
                f"Valor Unit.: R$ {item['valor_unitario'] or 0:.2f}"
            )
            c5 = ws_confronto.cell(row=row_num, column=5, value=item_ofertado_desc)
            c5.font = FONT_BODY
            c5.alignment = ALIGN_LEFT
            c5.border = BORDER_THIN
            
            # 6. Confronto IA
            c6 = ws_confronto.cell(row=row_num, column=6, value=item["confronto_ia"])
            c6.font = FONT_BODY
            c6.alignment = ALIGN_LEFT
            c6.border = BORDER_THIN
            
            # 7. Setor Competente (Manual)
            c7 = ws_confronto.cell(row=row_num, column=7)
            c7.font = FONT_BODY
            c7.alignment = ALIGN_CENTER
            c7.border = BORDER_THIN
            c7.fill = ICE_BLUE_FILL  # Cor clara para indicar preenchimento manual
            
            # 8. Parecer Técnico (Dropdown Manual - pré-preenchido com recomendação da IA)
            c8 = ws_confronto.cell(row=row_num, column=8, value=item["status_sugerido"])
            c8.font = FONT_BODY_BOLD
            c8.alignment = ALIGN_CENTER
            c8.border = BORDER_THIN
            c8.fill = ICE_BLUE_FILL  # Cor clara indicando interatividade
            
            # 9. Justificativa (Manual)
            c9 = ws_confronto.cell(row=row_num, column=9)
            c9.font = FONT_BODY
            c9.alignment = ALIGN_LEFT
            c9.border = BORDER_THIN
            c9.fill = ICE_BLUE_FILL  # Cor clara indicando preenchimento manual
            
            row_num += 1

    # Criação do Dropdown de Parecer Técnico nas células da coluna H
    dv = DataValidation(type="list", formula1='"Aceito,Não aceito,Pendência"', allow_blank=True)
    ws_confronto.add_data_validation(dv)
    
    # Aplica o dropdown para a coluna H nas linhas correspondentes aos dados
    if row_num > 2:
        dv.add(f"H2:H{row_num - 1}")
        
    # Ajuste de largura das colunas na aba de Confronto (Auto-fit com largura mínima)
    col_widths = {
        "A": 25,  # Fornecedor
        "B": 18,  # CNPJ
        "C": 28,  # PDF da Proposta
        "D": 22,  # Item TR
        "E": 35,  # Item Ofertado
        "F": 55,  # Confronto IA
        "G": 18,  # Setor Competente
        "H": 18,  # Parecer Técnico
        "I": 30   # Justificativa
    }
    
    for col_letter, width in col_widths.items():
        ws_confronto.column_dimensions[col_letter].width = width
        
    # Salva a planilha no caminho especificado
    wb.save(output_path)
    print(f"[Planilha Excel] Planilha salva com sucesso em: {output_path.name}")
